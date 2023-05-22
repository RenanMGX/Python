# -*- coding: utf-8 -*-
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from pandas import read_excel
import openpyxl
from openpyxl import load_workbook
import datetime
from num2words import num2words
import locale
import win32com.client as win32
from time import sleep
# definir o locale para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
#função que cria um pop up
def popup_completed(mensagem, e="Aviso"):
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(e, mensagem)
#função para escrever numero por extenso
def converter_para_extenso_e_monetario(numero):
    numero_inteiro, numero_fracionario = str(numero).split('.')
    numero_por_extenso = num2words(int(numero_inteiro), lang='pt_BR')
    numero_por_extenso2 = num2words(int(numero_fracionario[0:2]), lang='pt_BR')
    return f"{numero_por_extenso.capitalize()} reais e {numero_por_extenso2} centavos"

#função para alterar data para ultimo dia do mes anterior
def ultimo_dia_mes_anterior(data):
    # Subtrai um dia da data para obter o último dia do mês anterior
    primeiro_dia_mes_atual = datetime.date(data.year, data.month, 1)
    ultimo_dia_mes_anterior = primeiro_dia_mes_atual - datetime.timedelta(days=1)
    return ultimo_dia_mes_anterior
# pegar dia 20 do mesmo mes
def dia_vinte_do_mes(data):
    # Define a data do dia 20 do mês da data informada
    dia_vinte = datetime.datetime(data.year, data.month, 20).date()
    return dia_vinte
# função para passar a data por extenso
def converter_data(data, apenas_mes=False):
    data_obj = datetime.datetime.strptime(str(data), '%Y-%m-%d')
    ano = data_obj.strftime('%Y')
    
    # obter o nome do mês em português
    mes_extenso = data_obj.strftime('%B').capitalize()
    #mes_extenso = mes_extenso.encode('iso-8859-1').decode('utf-8')
    if mes_extenso == "Marã§o":
        mes_extenso = "Março"
    if apenas_mes == False:
        return f"{mes_extenso}/{ano}"
    else:
        return f"{mes_extenso}"
# função para colocar o perioro de 1 mes 
def periodo_um_mes(data):
    primeiro_dia_proximo_mes_altual = datetime.date(data.year, data.month, 1)
    ultimo_dia_mes = datetime.datetime(data.year, data.month, 1) + datetime.timedelta(days=32)
    primeiro_dia_proximo_mes = datetime.datetime(ultimo_dia_mes.year, ultimo_dia_mes.month, 1)
    ultimo_dia_mes_anterior = primeiro_dia_proximo_mes - datetime.timedelta(days=1)
    return f"{primeiro_dia_proximo_mes_altual.strftime('%d/%m/%Y')} a {ultimo_dia_mes_anterior.strftime('%d/%m/%Y')}"
# variavel com a data atual
data_atual = datetime.datetime.now().date()
#função procurar o arquivo Excel
def procurar_arquivo():
    root = tk.Tk()
    root.withdraw()
    try:
        file_path = filedialog.askopenfilename()
    except:
        return popup_completed("arquivo inexistente ou incorreto", e="Error")
    return file_path
# função para ler a folha 'RATEIO'
def ler_planilha(workbook, nome_sheet, colunas):
    sheet = workbook[nome_sheet]
    valores = []
    for row in sheet.iter_rows(min_row=8):
        linha = []
        for coluna in colunas:
            print(coluna)
            if isinstance(row[coluna].value, str):
                # Se a célula contém uma fórmula, avalia a fórmula e obtém o valor
                valor_formula = sheet.cell(row=row[coluna].row, column=row[coluna].column).value
                linha.append(valor_formula)
            else:
                # Se a célula contém um valor, adiciona o valor diretamente
                linha.append(row[coluna].value)
        valores.append(linha)
    resultado = []
    for x in valores:
        if x[0] != None:
            resultado.append(x)
    valores = []
    return resultado
# função para ler celula dentro de uma folha especifica
def ler_celula(workbook, sheetname, cell):
    sheet = workbook[sheetname]
    valor_celula = sheet[cell].value
    workbook.close()
    return valor_celula
# função para salvar os dados do dicionadio 'mapa' nas sheets
def salvar_excel(workbook, folha, valor, numero, data_temp):
   # Abrir arquivo Excel e selecionar a planilha desejada
    worksheet = workbook[folha]
    # Salvar dados nas células
    worksheet['B9'] = ultimo_dia_mes_anterior(data_atual)
    try:
        if data_temp.date() == dia_vinte_do_mes(data_atual):
            worksheet['D9'] = int(numero)
        else:
            worksheet['D9'] = int(numero) + 1
    except:
        worksheet['D9'] = int(numero) + 1
    worksheet['F9'] = valor
    worksheet['I9'] = dia_vinte_do_mes(data_atual)
    worksheet['B20'] = f"Valor correspondente ao reembolso de despesas administrativas da unidade matriz referente ao mês de {converter_data(str(data_atual))}."
    worksheet['B22'] = f"Período de {periodo_um_mes(data_atual)}"
    worksheet['D35'] = converter_para_extenso_e_monetario(valor)
    worksheet['T1'] = dia_vinte_do_mes(data_atual)
##################################################################
# função para salvar em PDF
def save_sheets_as_pdf(workbook, excel, sheetname):
    # Selecionando a aba desejada
    worksheet = workbook.Worksheets(sheetname)
    worksheet.Select()
    # Definindo as margens da página
    worksheet.PageSetup.LeftMargin = excel.InchesToPoints(0.25)
    worksheet.PageSetup.RightMargin = excel.InchesToPoints(0.25)
    worksheet.PageSetup.TopMargin = excel.InchesToPoints(0.75)
    worksheet.PageSetup.BottomMargin = excel.InchesToPoints(0.75)
    # Definindo o intervalo de impressão
    worksheet.PageSetup.PrintArea = 'A1:M47'
    caminho = f"\\\\server008\\G\\ARQ_PATRIMAR\\Setores\\Financas\\Planejamento_Financeiro\\CONTROLADORIA_ROTINAS\\Controles\\Reembolso de Despesas Administrativas\\{data_atual.year}\\{converter_data(data_atual, apenas_mes=True)}\\NDs\\"
    caminho = verificar_caminho(caminho)
    # Imprimindo a planilha como PDF
    pdf_path = f'{caminho}{sheetname}.pdf'
    worksheet.PrintOut(From=1, To=1, Copies=1, Preview=False, ActivePrinter='Microsoft Print to PDF', PrintToFile=True, Collate=False, PrToFileName=pdf_path)

    # Salvando e fechando a planilha
    workbook.Saved = True
    return caminho
# função verifica se o caminho existe se não existir tenta criar se n conseguir criar ele criar um padrao no C://
def verificar_caminho(caminho):
    while True:
        if os.path.exists(caminho):
            return caminho
        else:
            try:
                os.makedirs(caminho)
                return caminho
            except:
                caminho = f"C:\\Temporario_NDs\\{data_atual.year}\\{converter_data(data_atual, apenas_mes=True)}\\NDs\\"
################################### FIM FUNÇOES #########################
#variavel onde será salva os pep e os valores
arquivos = []
# o dicionario "mapa" vai salvar as informações na seguinte ordem de index
# 'PEP': 
# [
# [0] - sheet - 'empreendimendo';
# [1] - valor;
# [2] - numero documento;
# [3] - ultimo dia mes anterior;
# [4] - mes vencimanto;
# [5] - valor temp;
# [6] - data do vencimento temp
# ]
mapa = {}
#temporario

#salva o caminho na variavel
popup_completed(f"O Script será iniciado favor selecione a planilha", "Alerta")
planilha = procurar_arquivo()

try:
    read_excel(planilha, sheet_name='RATEIO', usecols='C,K', skiprows=8)
    workbook = openpyxl.load_workbook(planilha, data_only=True, keep_vba=True)
except ValueError:
    popup_completed("A planilha que você tentou abrir está incorreta", "Error")
    exit()
except PermissionError as permission:
    popup_completed(f"não foi possivel executar a planilha ela está aberta ou foi corrompida", "Error")
    with open("log.txt", "a") as log_error:
        log_error.write("\n" + str(data_atual) + ": \n" + str(permission))
        exit()
except Exception as e:
    with open("log.txt", "a") as log_error:
        log_error.write("\n" + str(data_atual) + ": \n" + str(e))
        exit()
# salva temporariamente o pep e o valor do Rateio
arquivos = ler_planilha(workbook, 'RATEIO', [2, 10])
folhas = workbook.sheetnames
# verifica se as sheets contem o numero do pep se tiver salva no dicionario 'mapa'
for pep in arquivos:
    for sheet in folhas:
        if pep[0] in sheet:
            mapa[pep[0]] = [sheet, pep[1]]
arquivos = []
# dentro de cada sheet ele salva os dados das celulas a beixo no dicionario 'mapa'
for key,valor in mapa.items():
    valor.append(ler_celula(workbook, valor[0], "D9"))
    valor.append(ler_celula(workbook, valor[0], "B9"))
    valor.append(ler_celula(workbook, valor[0], "I9"))
    valor.append(ler_celula(workbook, valor[0], "S1"))
    valor.append(ler_celula(workbook, valor[0], "T1"))

for key,item in mapa.items():
    salvar_excel(workbook, item[0], item[1], item[2], item[6])
workbook.save(planilha)
# fecha o arquivo excel
workbook.close()
excel = win32.DispatchEx('Excel.Application')
excel.Visible = False # tornando a janela do Excel invisível
excel.DisplayAlerts = False # desativando alertas
# Abrindo a planilha desejada
try:
    workbook_pdf = excel.Workbooks.Open(planilha)
except Exception as e:
    popup_completed(f"não foi possivel salvar os arquivos em PDF;\n \n {e}", "ERRRO")
    with open("log.txt", "a") as log_error:
        log_error.write("\n" + str(data_atual) + ": \n" + str(e))
        exit()
caminho_salvo = []
for key,sheet in mapa.items():
    caminho_salvo = save_sheets_as_pdf(workbook_pdf, excel, sheet[0])

workbook_pdf.Close()
# Finalizando o Excel
excel.Quit()
sleep(2)
popup_completed(f"script concluido com sucesso!\n os arquivos PDF foram salvos no caminho:\n {caminho_salvo}", "Concluido")
